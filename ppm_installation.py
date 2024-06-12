'''
History
2024-06-11 SAS - Download files from Sharepoint  00IT_PPM folder and store it to 00temp path
2024-06-11 SAS - Replace file name with the mail id
2024-06-12 SAS - Made a graph and send the mail 
'''
from shareplum import Office365, Site
from shareplum.site import Version
import sys
from datetime import date
import os
import json
import requests
import openpyxl
import re
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import win32com.client as win32
import time

ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
config_path = os.path.join(ROOT_DIR, 'config.json')

# Read json config file
with open(config_path) as config_file:
    config = json.load(config_file)
    config = config['share_point']

USERNAME = config['user']
PASSWORD = config['password']
SHAREPOINT_URL = config['url']
SHAREPOINT_SITE = config['site']

# Get today's date
today = date.today()
d1 = today.strftime("%Y-%m-%d")
print(d1)

# Authenticate and connect to the SharePoint site
authcookie = Office365(SHAREPOINT_URL, username=USERNAME, password=PASSWORD).GetCookies()
site = Site(SHAREPOINT_SITE, version=Version.v365, authcookie=authcookie)

# Specify the SharePoint folder
folder_name = f'Shared Documents/00IT_PPM'
folder = site.Folder(folder_name)

#'download_SP_files' definition is used to download sharepoint files and place it in a folder in local system
def download_SP_files(local_folder_path):
    # Ensure the local folder exists
    if not os.path.exists(local_folder_path):
        os.makedirs(local_folder_path)

    # List all files in the SharePoint folder
    sharepoint_files = folder.files

    # Download each file from the SharePoint folder to the local folder
    for sharepoint_file in sharepoint_files:
        file_name = sharepoint_file['Name']
        file_url = sharepoint_file['ServerRelativeUrl']
        local_file_path = os.path.join(local_folder_path, file_name)
        
        with requests.get(f"{SHAREPOINT_URL}{file_url}", cookies=authcookie, stream=True) as r:
            r.raise_for_status()
            with open(local_file_path, 'wb') as f:
                for chunk in r.iter_content(chunk_size=8192):
                    f.write(chunk)
        
        print(f"Downloaded {file_name} to {local_file_path}")

    print('Downloaded all files from SharePoint to local folder')


# 'rename_file_Names' is used to rename the files downloaded to a specific mail format
def rename_file_Names(local_folder_path):
    # Iterate through all files in the folder
    for filename in os.listdir(local_folder_path):
        if filename.endswith(".log") and filename !='.log':
            try:
                # Remove the specific part and replace underscores with '@'
                n = filename.replace('_PPMinstallation', '')
                new_filename = n.replace('_', '@')
                
                # Create the full file paths
                old_file = os.path.join(local_folder_path, filename)
                new_file = os.path.join(local_folder_path, new_filename)
                
                # Attempt to rename the file
                try:
                    os.rename(old_file, new_file)
                    print(f'Renamed: {filename} to {new_filename}')
                except FileExistsError:
                    # If the new file already exists, remove it and try renaming again
                    os.remove(new_file)
                    os.rename(old_file, new_file)
                    print(f'Renamed: {filename} to {new_filename}, replaced existing file.')
            except Exception as e:
                print(f'Error renaming file {filename}: {e}')

# 'find_last_occurrence' is used to find the last occurence of installed started on date and time , and it makes a excel file with it's details
def find_last_occurrence(local_folder_path,excel_path):
    if(os.path.exists(fr'{excel_path}')):
        os.remove(fr'{excel_path}')
    # Create a new Excel workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Installation Data'
    
    # Write the headers
    ws.append(["MAIL ID", "INSTALLED-ON","COMPANY"])
    ws.append(["Log Generated on ", d1 ,""])
    
    # Iterate through all files in the folder
    for filename in os.listdir(local_folder_path):
        if filename.endswith(".log") and filename !='.log':
            file_path = os.path.join(local_folder_path, filename)
            
            with open(file_path, 'r') as file:
                data = file.read()
                
                # Find the last occurrence of the sentence
                last_occurrence_index = data.rfind('Installation Started on ')
                mail_id = filename[:-4]
                # To get the company name from the mail id
                company_name = mail_id.split('@')[1]
                company_name = company_name.split(".")[0]
                if company_name == 'outlook':
                    company_name = 'Ekeren'
                if last_occurrence_index != -1:
                    # Find the end of the sentence (assuming it ends with a newline or is the end of the file)
                    end_index = data.find('\n', last_occurrence_index)
                    if end_index == -1:
                        end_index = len(data)
                    
                    # To get the date and time from the sentence
                    sentence = data[last_occurrence_index:end_index].replace('Installation Started on ', '')
                    
                    # Append the Mail , date and time , and company name  to the Excel sheet
                    ws.append([mail_id, sentence,company_name])
                else:
                    # Append the filename without '.log' and a message indicating no occurrence found
                    ws.append([mail_id, 'No occurrence found',company_name])
    
    # Save the workbook
    wb.save(excel_path)


# "Plot_graph" is used to plot the graph with the data obtained from the excel file and put the graph in the excel file 
def Plot_graph(file_path,image_path):
    # Load the Excel file
    # file_path = r'C:\00temp\PPM_installation\just\00Prog 2024-06-12 SAS installation_data.xlsx'
    sheet_name = 'Installation Data'  # Change this if your sheet has a different name

    # Read the data from the Excel file
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Extract the 'Company' column starting from row 3 (zero-indexed)
    company_data = df.iloc[2:, 2]

    # Count the occurrences of each value
    counts = company_data.value_counts()

    # Create the bar graph
    plt.figure(figsize=(10, 6))
    bars = plt.bar(counts.index, counts.values, color='skyblue',width=0.4)

    # Add annotations on top of the bars
    for bar in bars:
        yval = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2.0, yval, int(yval), va='bottom')  # va: vertical alignment

    # Add labels and title
    plt.xlabel('Company')
    plt.ylabel('Total')
    plt.title('Number of installation in each firm')
    plt.xticks(rotation=45)  # Rotate x-axis labels if necessary

    # Show the plot
    plt.tight_layout()  # Adjust layout to make room for labels
    # image_path = 'bar_graph.png'
    plt.savefig(image_path)
    # plt.show()


    workbook = load_workbook(file_path)

    # Create a new sheet for the image
    new_sheet_name = 'Chart'
    if new_sheet_name in workbook.sheetnames:
        sheet = workbook[new_sheet_name]
    else:
        sheet = workbook.create_sheet(title=new_sheet_name)

    # Add the image to the new sheet
    img = Image(image_path)
    img.anchor = 'C5'  # Position the image at cell
    sheet.add_image(img)

    # Save the workbook
    workbook.save(file_path)
    print(f"Image added to the Excel file in the sheet '{new_sheet_name}'")

# 'Draft_mail' is used to draft a mail with the image attached in a specific format
def Draft_mail(image_path):
    # Create the email body with HTML formatting
    html_body = f"""
    <html>
    <head></head>
    <body>
        <font size="2">
        <p>Betreft: <font color="blue">00PROG ALG - PLE - Stabiliteit - PE - Programs</font></p>
        <table border="1" cellpadding="5" cellspacing="0">
        <tr>
            <font size="2">
            <th bgcolor="blue">Soort</th>
            <th bgcolor="blue">PlanNr</th>
            <th bgcolor="blue">Titel</th>
            </font>
        </tr>
        <tr>
            <font size="2">
            <td>O3 PlanetPM</td>
            <td>PPM_Usage_Log_Graph</td>
            <td>Python PPM_Usage log Excel</td>
            </font>
        </tr>
        </table>
        <br>
        <p>OPM: Wij gebruiken een kleurcode - Gelieve alle documenten (plannen, nota's, verslagen, Schetsen, ...) steeds af te drukken in kleur + optie PDF printen te kiezen</p>
        <br>
        <p>Hello,</p>
        <p>Please find the attached PPM installation log as on {d1}. See graph below</p>
        <img src="cid:image1" width="650" height="400"><br>
        <p>Regards,<br>IT Team</p>
        </font>
    </body>
    </html>
    """

    # Create an Outlook application instance
    outlook = win32.Dispatch('outlook.application')

    # Create a new mail item
    mail = outlook.CreateItem(0)

    # Set email properties
    mail.To = 'receiver_email@example.com'
    mail.Subject = f'PPM Usage Log Graph {d1}'
    mail.HTMLBody = html_body

    # Attach the image
    attachment = mail.Attachments.Add(image_path)
    attachment.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", "image1")

    # To send mail drectly
    # mail.Send()
    # Display the mail
    mail.Display()
    sys.exit()


def select_folder():
    folder_selected = filedialog.askdirectory()
    if folder_selected:
        folder_path_entry.delete(0, tk.END)
        folder_path_entry.insert(0, folder_selected)


def on_download_click():
    folder_path = folder_path_entry.get()
    image_path = fr'{os.getcwd()}\bar_graph.png'
    
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    
    if os.path.isdir(folder_path):
        # To download files from SP and store it to a folder in local computer
        download_SP_files(folder_path)
        # To rename the files made in the folder
        rename_file_Names(folder_path)
        # To get the initial of the user
        file1 = open(r'C:\My Documents\01Data\IDPE.txt','r')
        line = file1.readlines()
        print(line[1])
        match = re.search(r'>(.*?)<', line[1])
        initial = "initial"
        if match:
            initial = match.group(1)

        # Call the function to find the last occurrence and write to Excel
        excel_path = fr'{os.getcwd()}\00Prog {d1} {initial} 00IT_PPM installation_data.xlsx' 
        find_last_occurrence(folder_path,excel_path) 
        time.sleep(0.1)
        Plot_graph(excel_path,image_path)
        time.sleep(0.1)
        Draft_mail(image_path)
        messagebox.showinfo("Success", "File Downloaded successfully!")
    else:
         messagebox.showerror("Error", "Invalid directory")
    root.destroy()

# Create the main window
root = tk.Tk()
root.title("Download 00IT_PPM files")

label = tk.Label(root, text="Destination file where files to be placed")
label.pack(side=tk.TOP, pady=10)

# Create and place the folder path entry and browse button
folder_path_entry = tk.Entry(root, width=50)
folder_path_entry.pack(padx=10, pady=10)

browse_button = tk.Button(root, text="Browse", command=select_folder)
browse_button.pack(padx=10, pady=5)

# Create and place the download button
download_button = tk.Button(root, text="Download Files", command=on_download_click)
download_button.pack(padx=10, pady=10)

CopyRight = tk.Label(root, text="Version 1.0 2024-06-12 Copyright@paradigm.in")
CopyRight.pack(side=tk.TOP, pady=10)

# Run the Tkinter event loop
root.mainloop()
